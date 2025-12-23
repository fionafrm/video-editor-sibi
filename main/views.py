from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from moviepy import concatenate_videoclips
from django.views.decorators.http import require_http_methods
from moviepy.video.io.VideoFileClip import VideoFileClip
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from .forms import CustomUserCreationForm, CustomPasswordResetForm, CustomSetPasswordForm, CustomPasswordChangeForm
import pandas as pd
import requests
import re
import openpyxl
from io import BytesIO
from django.contrib.auth.models import User
import os
import json
import csv
from .models import Video
import logging
from collections import defaultdict
import zipfile
from django.http import HttpResponse, FileResponse

logger = logging.getLogger(__name__)

"""DOWNLOAD FUNCTIONS"""

@login_required
def download_video(request, video_id):
    """Download video individual"""
    video = get_object_or_404(Video, id=video_id)
    
    if not video.file or not default_storage.exists(video.file.name):
        messages.error(request, f'File video tidak ditemukan: {video.title}')
        return redirect('main:landing_page')
    
    file_path = video.file.path
    
    try:
        response = FileResponse(
            open(file_path, 'rb'),
            content_type='video/mp4'
        )
        # Set filename untuk download
        safe_filename = "".join(c for c in video.title if c.isalnum() or c in (' ', '-', '_', '.')).strip()
        response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
        
        return response
    except Exception as e:
        messages.error(request, f'Error downloading video: {str(e)}')
        return redirect('main:landing_page')

@login_required
def download_folder_videos(request, folder_name):
    """Download semua video dalam folder sebagai ZIP"""
    videos = Video.objects.filter(folder_name=folder_name)
    
    if not videos.exists():
        messages.error(request, f'Tidak ada video di folder: {folder_name}')
        return redirect('main:landing_page')
    
    # Create ZIP in memory
    response = HttpResponse(content_type='application/zip')
    safe_folder_name = "".join(c for c in folder_name if c.isalnum() or c in (' ', '-', '_')).strip()
    response['Content-Disposition'] = f'attachment; filename="{safe_folder_name}.zip"'
    
    with zipfile.ZipFile(response, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for video in videos:
            if video.file and default_storage.exists(video.file.name):
                file_path = video.file.path
                
                # Add video to ZIP (hanya video, tanpa metadata)
                safe_title = "".join(c for c in video.title if c.isalnum() or c in (' ', '-', '_', '.')).strip()
                zip_file.write(file_path, f'{safe_folder_name}/{safe_title}')
    
    return response

@login_required
def download_all_videos(request):
    """Download semua video dari database sebagai ZIP (Admin only)"""
    if not request.user.is_superuser:
        messages.error(request, 'Unauthorized - Admin access required')
        return redirect('main:landing_page')
    
    # Get folder parameter from query string
    folder_filter = request.GET.get('folder', 'all')
    
    # Filter videos based on folder parameter
    if folder_filter and folder_filter != 'all':
        videos = Video.objects.filter(folder_name=folder_filter)
        zip_filename = f"{folder_filter}_videos.zip"
    else:
        videos = Video.objects.all()
        zip_filename = "all_videos_database.zip"
    
    if not videos.exists():
        messages.error(request, 'Tidak ada video di database')
        return redirect('main:landing_page')
    
    # Create ZIP in memory
    response = HttpResponse(content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    
    with zipfile.ZipFile(response, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Group by folder
        folder_dict = defaultdict(list)
        for video in videos:
            folder_dict[video.folder_name or 'uncategorized'].append(video)
        
        for folder_name, folder_videos in folder_dict.items():
            safe_folder = "".join(c for c in folder_name if c.isalnum() or c in (' ', '-', '_')).strip()
            
            for video in folder_videos:
                if video.file and default_storage.exists(video.file.name):
                    file_path = video.file.path
                    
                    # Add video to ZIP (hanya video, tanpa metadata)
                    safe_title = "".join(c for c in video.title if c.isalnum() or c in (' ', '-', '_', '.')).strip()
                    zip_file.write(file_path, f'{safe_folder}/{safe_title}')
    
    return response

@csrf_exempt
@login_required
def upload_video(request):
    """ Mengunggah video baru """
    if request.method == 'POST' and request.FILES.get('file'):
        video_file = request.FILES['file'].read()
        folder_name = request.POST.get('folder_name')
        video_title = request.POST.get('video_title')
        
        video_filename = f"{folder_name}_{video_title}.mp4"
        
        raw_video_folder = os.path.join('raw_videos', folder_name)
        video_folder = os.path.join('videos', folder_name)
        
        # Ensure the subdirectories exist
        os.makedirs(os.path.join(settings.MEDIA_ROOT, raw_video_folder), exist_ok=True)
        os.makedirs(os.path.join(settings.MEDIA_ROOT, video_folder), exist_ok=True)
        
        raw_video_path = os.path.join(raw_video_folder, video_filename)
        raw_video_file = default_storage.save(raw_video_path, ContentFile(video_file))
        
        video_path = os.path.join(video_folder, video_filename)
        video_file_new = default_storage.save(video_path, ContentFile(video_file))

        video = Video.objects.create(title=video_filename, file=video_file_new, folder_name=folder_name)

        return JsonResponse({'message': 'Video uploaded successfully', 'video_title': video_title})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@csrf_exempt
@login_required
def merge_videos(request, video_title):
    """ Menggabungkan video dengan video selanjutnya (berdasarkan urutan nama) """
    try:
        print(f"\n=== MERGE DEBUG untuk {video_title} ===")
        
        # Ambil video saat ini - gunakan filter().first() untuk menghindari error multiple objects
        video = Video.objects.filter(title=video_title).first()
        if not video:
            print(f"❌ Video tidak ditemukan: {video_title}")
            return JsonResponse({'error': f'Video tidak ditemukan: {video_title}'}, status=404)
        
        # Cek jika ada duplikasi
        duplicate_count = Video.objects.filter(title=video_title).count()
        if duplicate_count > 1:
            print(f"⚠️ WARNING: Ditemukan {duplicate_count} video dengan nama '{video_title}'!")
            print("Menggunakan video yang pertama ditemukan...")
        print(f"Video ditemukan: {video.title} (ID: {video.id})")
        print(f"Video file path: {video.file}")

        # Ekstrak sequence dari nama file, misal: "TVRI_SB_061119_0052.mp4"
        name_parts = video_title.replace('.mp4', '').split('_')
        print(f"Name parts: {name_parts}")
        
        current_sequence = int(name_parts[-1])
        base_title = '_'.join(name_parts[:-1])
        print(f"Current sequence: {current_sequence}")
        print(f"Base title: {base_title}")

        # Buat nama video berikutnya
        next_sequence = current_sequence + 1
        next_video_title = f"{base_title}_{next_sequence:04d}.mp4"
        print(f"Looking for next video: {next_video_title}")

        # Coba ambil video berikutnya dari database
        next_video = Video.objects.filter(title=next_video_title).first()
        
        if not next_video:
            print(f"❌ Video berikutnya tidak ditemukan: {next_video_title}")
            # Cek video apa saja yang ada dengan base_title yang sama
            similar_videos = Video.objects.filter(title__startswith=base_title).order_by('title')
            print(f"Video dengan base '{base_title}' yang ada:")
            for v in similar_videos:
                print(f"  - {v.title}")
            
            print(f"✅ Tidak ada video berikutnya, mengembalikan video asli: {video.title}")
            # Tidak ada video berikutnya, return video asli saja
            return JsonResponse({
                'message': 'Video asli (tidak ada video berikutnya untuk merge).',
                'merged_video_url': video.file.url,
                'is_single_video': True
            })
        
        # Cek duplikasi video berikutnya juga
        next_duplicate_count = Video.objects.filter(title=next_video_title).count()
        if next_duplicate_count > 1:
            print(f"⚠️ WARNING: Ditemukan {next_duplicate_count} video dengan nama '{next_video_title}'!")
            print("Menggunakan video yang pertama ditemukan...")
        
        print(f"✅ Video berikutnya ditemukan: {next_video.title} (ID: {next_video.id})")

        # Dapatkan path file kedua video
        path1 = video.file.path
        path2 = next_video.file.path
        
        print(f"Path video 1: {path1}")
        print(f"Path video 2: {path2}")
        print(f"File 1 exists: {os.path.exists(path1)}")
        print(f"File 2 exists: {os.path.exists(path2)}")

        if not os.path.exists(path1):
            print(f"❌ File pertama tidak ditemukan: {path1}")
            return JsonResponse({'error': f'File video pertama tidak ditemukan: {path1}'}, status=404)
        
        if not os.path.exists(path2):
            print(f"❌ File kedua tidak ditemukan: {path2}")
            return JsonResponse({'error': f'File video kedua tidak ditemukan: {path2}'}, status=404)

        logger.info(f"Merging {path1} + {path2}")
        print("🎬 Memulai proses merge dengan moviepy...")

        # Gabungkan menggunakan moviepy
        try:
            clip1 = VideoFileClip(path1)
            clip2 = VideoFileClip(path2)
            video_n_duration = clip1.duration  # Simpan durasi video n untuk marker
            print(f"Video 1 duration: {video_n_duration} seconds")
            print(f"Video 2 duration: {clip2.duration} seconds")
            
            merged_clip = concatenate_videoclips([clip1, clip2])
            print(f"Merged clip duration: {merged_clip.duration} seconds")

            # Simpan hasil merge
            merged_filename = f"merged_{video_title.replace('.mp4','')}_{next_video_title}"
            merged_path = os.path.join('edited_videos', merged_filename)
            
            print(f"Merged filename: {merged_filename}")
            print(f"Merged path: {merged_path}")
            
            # Buat folder edited_videos jika belum ada
            edited_dir = os.path.join(settings.MEDIA_ROOT, 'edited_videos')
            os.makedirs(edited_dir, exist_ok=True)
            print(f"Edited videos directory: {edited_dir}")
            
            # Path lengkap untuk file output
            full_merged_path = default_storage.path(merged_path)
            print(f"Full merged path: {full_merged_path}")
            
            print("📹 Writing merged video file...")
            merged_clip.write_videofile(full_merged_path, codec="libx264")
            print("✅ Video merge completed successfully")
            
            # Tutup clips untuk free memory
            clip1.close()
            clip2.close()
            merged_clip.close()

            # Simpan path dan durasi video n ke model
            video.merged_video_path = merged_path
            video.video_n_duration = video_n_duration  # Simpan durasi untuk marker
            video.save()
            print(f"✅ Database updated with merged_video_path: {merged_path}")
            print(f"✅ Video n duration (marker position): {video_n_duration} seconds")

            return JsonResponse({
                'message': 'Video berhasil digabung.',
                'merged_video_url': default_storage.url(merged_path),
                'video_n_duration': video_n_duration  # Kirim durasi untuk marker di timeline
            })
        
        except Exception as moviepy_error:
            print(f"❌ Error during moviepy processing: {moviepy_error}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': f'Error saat memproses video: {str(moviepy_error)}'}, status=500)

    except Exception as e:
        logger.error(f"Error in merging videos: {str(e)}")
        return JsonResponse({'error': f"Error saat menggabungkan video: {str(e)}"}, status=500)


@csrf_exempt
@login_required
def trim_video(request, video_title):
    """
    Algoritma Trim Video:
    1. Yang ditampilkan adalah merge dari video n dan n+1
    2. Saat trim:
       - Bagian yang di-keep (start_time sampai end_time) → overwrite video n
       - Bagian remainder (end_time sampai akhir merge) → overwrite video n+1
    3. Hapus merged_video_path dari video n
    4. Saat next, akan merge video n+1 (baru) dengan video n+2
    """
    if request.method == 'POST':
        video = get_object_or_404(Video, title=video_title)

        data = json.loads(request.body)
        start_time = data.get('start_time')
        end_time = data.get('end_time')

        if start_time is None or end_time is None:
            return JsonResponse({'error': 'Invalid start or end time'}, status=400)

        try:
            print(f"\n=== TRIM VIDEO {video_title} ===")
            print(f"Start time: {start_time}, End time: {end_time}")
            
            # Ekstrak informasi sequence dari nama file
            parts = video.title.replace('.mp4', '').split('_')
            current_seq = int(parts[-1])
            base = '_'.join(parts[:-1])
            
            # Cari video berikutnya
            next_title = f"{base}_{current_seq+1:04d}.mp4"
            next_video = Video.objects.filter(title=next_title).first()
            
            # Gunakan merged video jika ada, kalau tidak ada gunakan video asli
            if video.merged_video_path and default_storage.exists(video.merged_video_path):
                # Gunakan merged video (n+n+1)
                source_path = default_storage.path(video.merged_video_path)
                print(f"Using merged video as source: {source_path}")
            else:
                # Fallback ke video asli di raw_videos
                source_path = os.path.join(settings.MEDIA_ROOT, 'raw_videos', video.folder_name, video.title)
                print(f"Using raw video as source: {source_path}")
            
            if not os.path.exists(source_path):
                return JsonResponse({'error': f'Source video not found: {source_path}'}, status=404)
            
            # Load video source
            source_clip = VideoFileClip(source_path)
            source_duration = source_clip.duration
            print(f"Source video duration: {source_duration} seconds")

            # Validasi waktu
            if end_time > source_duration:
                end_time = source_duration - 0.01
                print(f"Adjusted end_time to {end_time}")

            if start_time >= end_time:
                source_clip.close()
                return JsonResponse({'error': 'Invalid time range'}, status=400)

            # BAGIAN 1: Trim video (start_time sampai end_time) → overwrite video n
            trimmed_clip = source_clip.subclipped(start_time, end_time)
            final_output_path = video.file.path
            print(f"Writing trimmed video to video n: {final_output_path}")
            trimmed_clip.write_videofile(final_output_path, codec="libx264", audio_codec="aac")
            trimmed_clip.close()
            
            # BAGIAN 2: Remainder (end_time sampai akhir) → overwrite video n+1
            if next_video and end_time < source_duration:
                try:
                    remainder_clip = source_clip.subclipped(end_time, source_duration)
                    next_video_path = next_video.file.path
                    print(f"Writing remainder to video n+1: {next_video_path}")
                    remainder_clip.write_videofile(next_video_path, codec="libx264", audio_codec="aac")
                    remainder_clip.close()
                    
                    # Hapus merged_video_path dari video n+1 jika ada
                    if next_video.merged_video_path:
                        print(f"Clearing merged_video_path from video n+1")
                        next_video.merged_video_path = None
                        next_video.save()
                    
                    print(f"✅ Successfully overwritten video n+1 with remainder")
                except Exception as remainder_error:
                    logger.error(f"Error saving remainder to next video: {str(remainder_error)}")
                    import traceback
                    traceback.print_exc()
            else:
                if not next_video:
                    print(f"ℹ️ No next video found ({next_title})")
                if end_time >= source_duration:
                    print(f"ℹ️ No remainder (trim ends at video end)")
            
            # Clean up
            source_clip.close()
            
            # BAGIAN 3: Hapus merged_video_path dari video n agar next time merge ulang
            if video.merged_video_path:
                print(f"Clearing merged_video_path from video n")
                video.merged_video_path = None
                video.save()

            print(f"✅ Trim completed successfully")
            return JsonResponse({
                'message': 'Video trimmed and saved successfully.',
                'trimmed_video_url': video.file.url
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@login_required
def save_transcript(request, video_title):
    if request.method == 'POST':
        data = json.loads(request.body)
        video = get_object_or_404(Video, title=video_title)

        transcript_alignment = data.get('transcript_alignment')
        sibi_sentence = data.get('sibi_sentence')
        potential_problem = data.get('potential_problem')
        comment = data.get('comment')

        if transcript_alignment is not None:
            video.transcript_alignment = transcript_alignment
        if sibi_sentence is not None:
            video.sibi_sentence = sibi_sentence
        if potential_problem is not None:
            video.potential_problem = potential_problem
        if comment is not None:
            video.comment = comment

        video.is_annotated = True
        video.annotated_by = request.user
        video.save()

        return JsonResponse({'message': 'Data disimpan'})
    return JsonResponse({'error': 'Invalid request'}, status=400)


@csrf_exempt
@login_required
def get_video_details(request, video_title):
    """ Mengambil informasi video, transkripnya, dan komennya """
    video = get_object_or_404(Video, title=video_title)
    return JsonResponse({
        'video_url': video.file.url ,
        'transcript': video.transcript,
        'comment': video.comment
    })

@csrf_exempt
@login_required
def get_merged_video(request, video_title):
    print(f"\n=== GET_MERGED_VIDEO DEBUG untuk {video_title} ===")
    
    try:
        # Gunakan filter().first() untuk menghindari error multiple objects
        video = Video.objects.filter(title=video_title).first()
        if not video:
            print(f"❌ Video tidak ditemukan: {video_title}")
            return JsonResponse({'error': f'Video tidak ditemukan: {video_title}'}, status=404)
        
        # Cek jika ada duplikasi
        duplicate_count = Video.objects.filter(title=video_title).count()
        if duplicate_count > 1:
            print(f"⚠️ WARNING: Ditemukan {duplicate_count} video dengan nama yang sama!")
            print("Menggunakan video yang pertama ditemukan...")
        print(f"Video ditemukan: {video.title} (ID: {video.id})")
        print(f"Merged video path: {video.merged_video_path}")
        
        # Cek apakah file merge sudah ada
        if video.merged_video_path:
            file_exists = default_storage.exists(video.merged_video_path)
            print(f"Merged file exists: {file_exists}")
            if file_exists:
                full_path = default_storage.path(video.merged_video_path)
                print(f"Full path: {full_path}")
                print(f"File size: {os.path.getsize(full_path)} bytes")
        
        if video.merged_video_path and default_storage.exists(video.merged_video_path):
            merged_video_url = default_storage.url(video.merged_video_path)
            print(f"✅ Returning existing merged video: {merged_video_url}")
            
            # Ambil durasi video n untuk marker
            video_n_duration = getattr(video, 'video_n_duration', None)
            
            # Jika tidak ada di database, hitung dari file asli
            if video_n_duration is None:
                try:
                    video_path = video.file.path
                    if os.path.exists(video_path):
                        temp_clip = VideoFileClip(video_path)
                        video_n_duration = temp_clip.duration
                        temp_clip.close()
                        print(f"✅ Calculated video_n_duration: {video_n_duration} seconds")
                except Exception as e:
                    print(f"⚠️ Could not calculate video_n_duration: {e}")
                    video_n_duration = None
            
            return JsonResponse({
                'merged_video_url': merged_video_url,
                'transcript': video.transcript,
                'comment': video.comment,
                'video_n_duration': video_n_duration  # Kirim durasi untuk marker
            })

        print("❌ No existing merged video, attempting to create new one...")
        # Jika belum ada hasil merge, lakukan merge dulu
        merge_response = merge_videos(request, video_title)
        print(f"Merge response status: {merge_response.status_code}")

        # Kalau sukses (status 200), process response
        if merge_response.status_code == 200:
            # Parse response untuk cek apakah ini single video atau merged video
            response_data = json.loads(merge_response.content.decode('utf-8'))
            
            if response_data.get('is_single_video'):
                # Video tunggal (tidak ada video berikutnya)
                print(f"✅ Single video returned: {response_data.get('merged_video_url')}")
                return JsonResponse({
                    'merged_video_url': response_data.get('merged_video_url'),
                    'transcript': video.transcript,
                    'comment': video.comment,
                    'is_single_video': True,
                    'message': 'Video asli (tidak ada video berikutnya untuk merge)',
                    'video_n_duration': None  # Tidak ada marker untuk single video
                })
            else:
                # Video berhasil di-merge
                video.refresh_from_db()  # Pastikan ambil data terbaru dari DB
                print(f"After merge - merged_video_path: {video.merged_video_path}")
                if video.merged_video_path:
                    merged_video_url = default_storage.url(video.merged_video_path)
                    video_n_duration = response_data.get('video_n_duration') or getattr(video, 'video_n_duration', None)
                    print(f"✅ Merge successful, returning: {merged_video_url}")
                    print(f"✅ Video n duration (marker): {video_n_duration} seconds")
                    return JsonResponse({
                        'merged_video_url': merged_video_url,
                        'transcript': video.transcript,
                        'comment': video.comment,
                        'is_single_video': False,
                        'video_n_duration': video_n_duration  # Kirim durasi untuk marker
                    })
                else:
                    print("❌ Merge reported success but no merged_video_path saved")
                    return JsonResponse({'error': 'Merge completed but path not saved'}, status=500)

        # Kalau merge gagal, return responsenya langsung
        print("❌ Merge failed, returning error response")
        return merge_response
        
    except Exception as e:
        print(f"❌ Exception in get_merged_video: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)

@csrf_exempt
@login_required
def get_next_video_status(request, folder_name, current_title):
    videos = Video.objects.filter(folder_name=folder_name, is_annotated=False).order_by('title')

    for video in videos:
        if video.title != current_title:
            return JsonResponse({'has_next': True})
    
    return JsonResponse({'has_next': False})


@csrf_exempt
@login_required
def upload_transcript_csv(request):
    """ Mengunggah transkrip dari file CSV """
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        file_path = default_storage.save(f"transcripts/{file.name}", file)

        with default_storage.open(file_path, 'r') as f:
            reader = csv.reader(f)
            next(reader)  # Skip header

            for row in reader:
                video_id_title, transcript = row[0], row[1]
                last_4_digits = f"{int(video_id_title):04d}"
                try:
                    video = get_object_or_404(Video, title__endswith=f"{last_4_digits}.mp4")
                    video.transcript = transcript
                    video.save()
                except Exception as e:
                    # Debugging: Catch any exceptions and print the error
                    print(f"Error for video_id: {video_id_title} - {str(e)}")

        return JsonResponse({'message': 'Transcript uploaded successfully.'})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@csrf_exempt
@login_required
def upload_file(request):
    """Mengunggah file Excel dengan link Google Drive di hyperlink kolom A"""
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        tmp_path = default_storage.save(f"temp/{file.name}", file)

        # Logic sederhana: gunakan file yang sudah ada jika valid, atau download jika tidak ada
        print("Mode: Otomatis - gunakan file lokal jika ada, atau download jika tidak ada")

        try:

            file_path = default_storage.path(tmp_path)

            # Baca isi tabel ke DataFrame
            df = pd.read_excel(file_path, engine='openpyxl')
            
            print(f"DataFrame shape: {df.shape}")
            print(f"DataFrame columns: {list(df.columns)}")
            print("First few rows:")
            print(df.head())
            
            # Cek apakah kolom 'Nama Data' ada
            if 'Nama Data' not in df.columns:
                available_cols = list(df.columns)
                print("ERROR: Kolom 'Nama Data' tidak ditemukan!")
                print(f"Kolom yang tersedia: {available_cols}")
                return JsonResponse({'error': f'Kolom "Nama Data" tidak ditemukan. Kolom tersedia: {available_cols}'}, status=400)

            # Baca workbook untuk ambil hyperlink kolom A
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            print(f"Worksheet loaded: {ws.title}")
            print(f"Total rows in worksheet: {ws.max_row}")
            print(f"Total columns in worksheet: {ws.max_column}")

            # Ambil hyperlink dari kolom A (A2, A3, ...)
            drive_links = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):  # skip header
                cell = row[0]  # kolom A
                
                # Cek hyperlink dulu, lalu cell value
                if cell.hyperlink:
                    link = cell.hyperlink.target
                    print(f"Row {row_num}: Found hyperlink = '{link}'")
                elif cell.value:
                    link = str(cell.value).strip()
                    print(f"Row {row_num}: Found cell value = '{link}'")
                else:
                    link = ""
                    print(f"Row {row_num}: No link found (empty cell)")
                
                drive_links.append(link)

            count = 1
            print(f"Total baris data: {len(df)}")
            print(f"Total drive links: {len(drive_links)}")
            
            for idx, row in df.iterrows():
                try:
                    # Debugging informasi baris
                    print(f"\n=== Memproses Baris {idx+2} ===")
                    
                    # Cek apakah idx dalam range drive_links
                    if idx >= len(drive_links):
                        print(f"[SKIP] Baris {idx+2}: Index melebihi jumlah drive_links ({len(drive_links)})")
                        continue
                    
                    link = str(drive_links[idx]).strip() if drive_links[idx] else ""
                    
                    # Cek kolom 'Nama Data'
                    nama_data_col = row.get('Nama Data')
                    if nama_data_col is None:
                        print(f"[SKIP] Baris {idx+2}: Kolom 'Nama Data' tidak ditemukan")
                        print(f"Available columns: {list(row.keys())}")
                        continue
                    
                    video_title_raw = str(nama_data_col).strip()
                    
                    # Detail debugging
                    print(f"Link: '{link}'")
                    print(f"Video title raw: '{video_title_raw}'")
                    print(f"Link valid: {'drive.google.com' in link}")
                    print(f"Video title valid: {video_title_raw not in ['', 'nan', 'None']}")

                    # Validasi dan normalisasi link
                    if not link or link in ['nan', 'None', '']:
                        print(f"[SKIP] Baris {idx+2}: Link kosong atau invalid. Link: '{link}'")
                        continue
                    
                    # Coba normalisasi link jika tidak lengkap
                    if 'drive.google.com' not in link:
                        # Jika hanya file ID, buat URL lengkap
                        if re.match(r'^[a-zA-Z0-9_-]{25,}$', link.strip()):
                            link = f"https://drive.google.com/file/d/{link.strip()}/view"
                            print(f"Link dinormalisasi menjadi: {link}")
                        else:
                            print(f"[SKIP] Baris {idx+2}: Bukan link Google Drive atau file ID. Link: '{link}'")
                            continue
                        
                    if not video_title_raw or video_title_raw in ['nan', 'None', '']:
                        print(f"[SKIP] Baris {idx+2}: Video title kosong atau invalid. Title: '{video_title_raw}'")
                        continue

                    video_title = f"{video_title_raw}.mp4"
                    folder_name = "_".join(video_title_raw.split("_")[:-1])

                    # Ambil metadata lain dengan debugging
                    automated_transcript = str(row.get('Transkripsi Suara secara Otomatis oleh Sistem', '')).strip()
                    transcript_alignment = str(row.get('Penyelarasan Suara/Teks Transkripsi dan Gerakan Bahasa Isyarat', '')).strip()
                    sibi_sentence = str(row.get('Kalimat yang Diperagakan', '')).strip()
                    potential_problem = str(row.get('Potensi Masalah', '')).strip()
                    comment = str(row.get('Keterangan Annotator', '')).strip()
                    username = str(row.get('Nama Annotator', '')).strip()
                    
                    print("Metadata extracted:")
                    print(f"- Automated transcript: {len(automated_transcript)} chars")
                    print(f"- Transcript alignment: {len(transcript_alignment)} chars")
                    print(f"- SIBI sentence: {len(sibi_sentence)} chars")
                    print(f"- Potential problem: {len(potential_problem)} chars")
                    print(f"- Comment: {len(comment)} chars")
                    print(f"- Username: '{username}'")
                    
                    # Parsing kolom Hasil Alignment (NEW) dengan debugging
                    hasil_alignment_raw = row.get('Hasil Alignment (NEW)')
                    print(f"Hasil Alignment (NEW) raw value: '{hasil_alignment_raw}' (type: {type(hasil_alignment_raw)})")
                    
                    if hasil_alignment_raw is None or hasil_alignment_raw == '':
                        is_annotated = False
                        print("is_annotated = False (kolom kosong)")
                    else:
                        # Konversi ke string dan normalisasi
                        hasil_alignment_str = str(hasil_alignment_raw).strip().lower()
                        print(f"Hasil Alignment (NEW) normalized: '{hasil_alignment_str}'")
                        
                        # Cek berbagai format yang mungkin
                        if hasil_alignment_str in ['1', '1.0', 'true', 'yes', 'ya', 'sudah']:
                            is_annotated = True
                            print("is_annotated = True (nilai positif terdeteksi)")
                        elif hasil_alignment_str in ['0', '0.0', 'false', 'no', 'tidak', 'belum', '']:
                            is_annotated = False
                            print("is_annotated = False (nilai negatif terdeteksi)")
                        else:
                            # Coba parsing sebagai angka
                            try:
                                numeric_value = float(hasil_alignment_str)
                                is_annotated = numeric_value > 0
                                print(f"is_annotated = {is_annotated} (parsed as number: {numeric_value})")
                            except ValueError:
                                is_annotated = False
                                print(f"is_annotated = False (tidak bisa parsing: '{hasil_alignment_str}')")
                    
                    print(f"Final is_annotated value: {is_annotated}")

                    # Parsing username dengan debugging
                    print(f"Username raw: '{username}'")
                    try:
                        if username and username not in ['', 'nan', 'None']:
                            user = User.objects.get(username=username)
                            print(f"User found: {user.username} (ID: {user.id})")
                        else:
                            user = None
                            print("No username provided, user = None")
                    except User.DoesNotExist:
                        user = None
                        print(f"User with username '{username}' not found in database")

                    # Ambil file ID dari link - Support berbagai format Google Drive
                    file_id = None
                    
                    # Pattern 1: /d/FILE_ID/view atau /d/FILE_ID
                    match = re.search(r'/d/([a-zA-Z0-9_-]+)', link)
                    if match:
                        file_id = match.group(1)
                        print(f"File ID extracted (pattern 1): {file_id}")
                    else:
                        # Pattern 2: id=FILE_ID
                        match = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', link)
                        if match:
                            file_id = match.group(1)
                            print(f"File ID extracted (pattern 2): {file_id}")
                        else:
                            # Pattern 3: Link berakhir dengan file ID
                            match = re.search(r'([a-zA-Z0-9_-]{25,})/?$', link)
                            if match:
                                file_id = match.group(1)
                                print(f"File ID extracted (pattern 3): {file_id}")
                    
                    if not file_id:
                        print(f"[SKIP] Baris {idx+2}: Gagal ekstrak file ID dari link: {link}")
                        print("Supported formats:")
                        print("- https://drive.google.com/file/d/FILE_ID/view")
                        print("- https://drive.google.com/open?id=FILE_ID")
                        print("- https://drive.google.com/file/d/FILE_ID")
                        continue
                    
                    download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
                    print(f"Download URL: {download_url}")

                    # Siapkan path untuk file sebelum melakukan pengecekan
                    videos_dir = os.path.join('videos', folder_name)
                    raw_dir = os.path.join('raw_videos', folder_name)
                    final_video_path = os.path.join(videos_dir, video_title)
                    raw_video_path = os.path.join(raw_dir, video_title)
                    
                    # Path absolut untuk pengecekan file di local storage
                    final_full_path = os.path.join(settings.MEDIA_ROOT, final_video_path)
                    raw_full_path = os.path.join(settings.MEDIA_ROOT, raw_video_path)
                    
                    # Fungsi untuk validasi file video
                    def is_valid_video_file(file_path):
                        """Validasi apakah file adalah video yang valid dan tidak corrupt"""
                        try:
                            if not os.path.exists(file_path):
                                return False
                            
                            # Cek ukuran file minimal (100KB untuk video)
                            file_size = os.path.getsize(file_path)
                            if file_size < 100000:  # 100KB minimum
                                print(f"File terlalu kecil: {file_size} bytes")
                                return False
                            
                            # Cek ekstensi file
                            if not file_path.lower().endswith(('.mp4', '.avi', '.mov', '.mkv')):
                                print(f"Ekstensi file tidak valid: {file_path}")
                                return False
                            
                            print(f"File valid: {file_path} ({file_size} bytes)")
                            return True
                        except Exception as e:
                            print(f"Error validating file {file_path}: {e}")
                            return False
                    
                    # Cek apakah file sudah ada di local storage
                    file_exists_in_final = is_valid_video_file(final_full_path)
                    file_exists_in_raw = is_valid_video_file(raw_full_path)
                    file_exists_locally = file_exists_in_final or file_exists_in_raw
                    
                    if file_exists_locally:
                        existing_path = final_full_path if file_exists_in_final else raw_full_path
                        file_size = os.path.getsize(existing_path)
                        print(f"📁 File sudah ada di local storage: {existing_path}")
                        print(f"File size: {file_size:,} bytes ({file_size/1024/1024:.1f} MB)")
                        print(f"[SKIP] Baris {idx+2}: File sudah ada dan valid, melewati download")
                        
                        # Tetap buat atau update record di database jika diperlukan
                        existing_videos = Video.objects.filter(title=video_title, folder_name=folder_name)
                        existing_video = existing_videos.first()
                        
                        # Cek jika ada duplikasi dan hapus yang extra
                        if existing_videos.count() > 1:
                            print(f"⚠️ WARNING: Ditemukan {existing_videos.count()} video duplikat dengan nama {video_title}")
                            # Hapus duplikasi, sisakan yang pertama
                            for duplicate in existing_videos[1:]:
                                print(f"Menghapus duplikasi video ID: {duplicate.id}")
                                duplicate.delete()
                            existing_video = existing_videos.first()
                        
                        if not existing_video:
                            print("File ada di local tapi tidak ada record di database, membuat record baru...")
                            # Gunakan path file yang sudah ada
                            relative_path = final_video_path if file_exists_in_final else raw_video_path
                            video_obj = Video.objects.create(
                                title=video_title,
                                folder_name=folder_name,
                                file=relative_path,
                                automated_transcript=automated_transcript,
                                transcript_alignment=transcript_alignment,
                                sibi_sentence=sibi_sentence,
                                potential_problem=potential_problem,
                                comment=comment,
                                annotated_by=user,
                                is_annotated=is_annotated
                            )
                            print(f"✅ Record database dibuat untuk file yang sudah ada (ID: {video_obj.id})")
                            count += 1
                        else:
                            # Update metadata dari Excel meskipun file sudah ada
                            print(f"Record database sudah ada (ID: {existing_video.id}), update metadata dari Excel...")
                            existing_video.automated_transcript = automated_transcript
                            existing_video.transcript_alignment = transcript_alignment
                            existing_video.sibi_sentence = sibi_sentence
                            existing_video.potential_problem = potential_problem
                            existing_video.comment = comment
                            if user:
                                existing_video.annotated_by = user
                            existing_video.is_annotated = is_annotated
                            existing_video.save()
                            print(f"✅ Metadata berhasil diperbarui (ID: {existing_video.id})")
                            count += 1
                        continue

                    # Buat folder jika belum ada
                    os.makedirs(os.path.join(settings.MEDIA_ROOT, videos_dir), exist_ok=True)
                    os.makedirs(os.path.join(settings.MEDIA_ROOT, raw_dir), exist_ok=True)

                    # Unduh video dengan handling untuk large files
                    print(f"Attempting to download from: {download_url}")
                    response = requests.get(download_url, stream=True)
                    
                    if response.status_code != 200:
                        print(f"[SKIP] Baris {idx+2}: Gagal download video (status {response.status_code})")
                        print(f"Response headers: {dict(response.headers)}")
                        continue
                    
                    # Cek apakah response berisi HTML (Google Drive confirmation page)
                    content_type = response.headers.get('content-type', '')
                    if 'text/html' in content_type:
                        print(f"[SKIP] Baris {idx+2}: Response berisi HTML, mungkin perlu konfirmasi download")
                        print("Kemungkinan file terlalu besar atau memerlukan konfirmasi manual")
                        
                        # Coba ekstrak confirmation URL jika ada
                        html_content = response.text
                        confirm_match = re.search(r'action="([^"]*)" method="post"', html_content)
                        if confirm_match:
                            confirm_url = confirm_match.group(1).replace('&amp;', '&')
                            print(f"Found confirmation URL: {confirm_url}")
                            # Coba download dengan confirmation
                            confirm_response = requests.get(f"https://drive.google.com{confirm_url}", stream=True)
                            if confirm_response.status_code == 200 and 'video' in confirm_response.headers.get('content-type', ''):
                                response = confirm_response
                            else:
                                print("[SKIP] Confirmation download juga gagal")
                                continue
                        else:
                            continue
                    
                    # Verifikasi content-type untuk memastikan ini file video
                    if response.status_code == 200:
                        content_type = response.headers.get('content-type', '')
                        content_length = response.headers.get('content-length', 'Unknown')
                        print(f"Content-Type: {content_type}, Content-Length: {content_length}")
                        
                        # Baca content untuk file yang berhasil didownload
                        video_content = response.content
                        if len(video_content) < 1000:  # File terlalu kecil, kemungkinan bukan video
                            print(f"[SKIP] Baris {idx+2}: File terlalu kecil ({len(video_content)} bytes), kemungkinan bukan video")
                            continue

                    # Path sudah disiapkan sebelumnya, langsung simpan
                    print(f"Saving video to: {final_video_path}")
                    print(f"Video size: {len(video_content):,} bytes ({len(video_content)/1024/1024:.1f} MB)")

                    final_path = default_storage.save(final_video_path, ContentFile(video_content))
                    default_storage.save(raw_video_path, ContentFile(video_content))

                    # Simpan ke DB
                    print("Creating Video record:")
                    print(f"- Title: {video_title}")
                    print(f"- Folder: {folder_name}")
                    print(f"- Annotated by: {user.username if user else 'None'}")
                    print(f"- Is annotated: {is_annotated}")
                    
                    video_obj = Video.objects.create(
                        title=video_title,
                        folder_name=folder_name,
                        file=final_path,
                        automated_transcript=automated_transcript,
                        transcript_alignment=transcript_alignment,
                        sibi_sentence=sibi_sentence,
                        potential_problem=potential_problem,
                        comment=comment,
                        annotated_by=user,
                        is_annotated=is_annotated
                    )
                    
                    print(f"✅ SUCCESS: Video {video_title} berhasil diproses dan disimpan (ID: {video_obj.id})")
                    count += 1

                except Exception as e:
                    print(f"[ERROR] Baris {idx+2}: {str(e)}")
                    import traceback
                    traceback.print_exc()

            print("\n=== SUMMARY ===")
            print(f"Total baris diproses: {len(df)}")
            print(f"Video berhasil diupload: {count-1}")
            return JsonResponse({'message': f'Upload selesai. {count-1} video diproses dari {len(df)} baris.'})

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=400)

@csrf_exempt
@login_required
def cleanup_duplicate_videos(request):
    """Fungsi untuk membersihkan video duplikat di database"""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Unauthorized - Admin access required'}, status=403)
    
    try:
        from collections import defaultdict
        
        # Grup video berdasarkan title dan folder_name
        video_groups = defaultdict(list)
        all_videos = Video.objects.all()
        
        for video in all_videos:
            key = (video.title, video.folder_name)
            video_groups[key].append(video)
        
        cleanup_summary = []
        total_deleted = 0
        
        for (title, folder), videos in video_groups.items():
            if len(videos) > 1:
                # Ada duplikasi
                print(f"Found {len(videos)} duplicates for {title} in folder {folder}")
                
                # Simpan yang pertama, hapus sisanya
                keep_video = videos[0]
                duplicates = videos[1:]
                
                for duplicate in duplicates:
                    print(f"Deleting duplicate video ID: {duplicate.id}")
                    duplicate.delete()
                    total_deleted += 1
                
                cleanup_summary.append({
                    'title': title,
                    'folder': folder,
                    'duplicates_found': len(videos),
                    'duplicates_deleted': len(duplicates),
                    'kept_video_id': keep_video.id
                })
        
        return JsonResponse({
            'message': f'Cleanup completed. {total_deleted} duplicate videos deleted.',
            'total_deleted': total_deleted,
            'cleanup_details': cleanup_summary
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
def debug_video_merge(request, video_title):
    """Debug endpoint untuk mengecek masalah merge"""
    try:
        print(f"\n=== DEBUG MERGE UNTUK {video_title} ===")
        
        # Cek video yang diminta
        video = Video.objects.filter(title=video_title).first()
        if not video:
            return JsonResponse({'error': f'Video tidak ditemukan: {video_title}'}, status=404)
        
        # Cek duplikasi
        duplicate_count = Video.objects.filter(title=video_title).count()
        print(f"✅ Video ditemukan: {video.title} (ID: {video.id})")
        if duplicate_count > 1:
            print(f"⚠️ WARNING: Ditemukan {duplicate_count} video duplikat!")
        
        # Parse nama file
        name_parts = video_title.replace('.mp4', '').split('_')
        current_sequence = int(name_parts[-1])
        base_title = '_'.join(name_parts[:-1])
        
        # Cari video berikutnya
        next_sequence = current_sequence + 1
        next_video_title = f"{base_title}_{next_sequence:04d}.mp4"
        
        print(f"Looking for next video: {next_video_title}")
        
        next_video = Video.objects.filter(title=next_video_title).first()
        
        # Cek semua video dengan base yang sama
        similar_videos = Video.objects.filter(title__startswith=base_title).order_by('title')
        video_list = [v.title for v in similar_videos]
        
        debug_info = {
            'current_video': {
                'title': video.title,
                'id': video.id,
                'file_exists': os.path.exists(video.file.path) if video.file else False,
                'merged_path': video.merged_video_path,
                'merged_exists': default_storage.exists(video.merged_video_path) if video.merged_video_path else False
            },
            'parsing': {
                'name_parts': name_parts,
                'current_sequence': current_sequence,
                'base_title': base_title,
                'next_sequence': next_sequence,
                'next_video_title': next_video_title
            },
            'next_video': {
                'found': next_video is not None,
                'title': next_video.title if next_video else None,
                'id': next_video.id if next_video else None,
                'file_exists': os.path.exists(next_video.file.path) if next_video and next_video.file else False
            },
            'similar_videos': video_list
        }
        
        return JsonResponse(debug_info)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["DELETE"])
@login_required
def delete_video(request, video_title):
    """ Menghapus video beserta file fisiknya """
    video = get_object_or_404(Video, title=video_title)

    # Hapus file video dari storage
    if video.file and video.file.storage.exists(video.file.name):
        video.file.delete()

    # Hapus file transkrip jika ada
    transcript_path = os.path.join(settings.MEDIA_ROOT, 'transcripts', f'video_{video_title}.txt')
    if os.path.exists(transcript_path):
        os.remove(transcript_path)

    # Hapus record di database
    video.delete()

    return JsonResponse({'message': f'Video dengan ID {video_title} berhasil dihapus'})


@csrf_exempt
@login_required
def search_videos(request, folder_name):
    # Ambil semua video dalam folder, urutkan berdasarkan nama
    videos = Video.objects.filter(folder_name=folder_name).order_by('title')

    for video in videos:
        if not video.is_annotated:
            return redirect('main:video_editor', video_title=video.title)

    # Jika semua sudah dianotasi, redirect ke folder_page
    messages.info(request, 'Semua video sudah dianotasi.')
    return redirect('main:folder_page', folder_name=folder_name)

@csrf_exempt
@login_required
def get_previous_video(request, folder_name):
    user = request.user
    # Cari video yang sudah dianotasi oleh user dan urutkan mundur
    videos = Video.objects.filter(folder_name=folder_name, is_annotated=True, annotated_by=user).order_by('-created_at')

    if videos.exists():
        return JsonResponse({'previous_video_title': videos.first().title})
    return JsonResponse({'previous_video_title': None})

"""AUTENTIKASI"""

def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        referral_code = request.POST.get('referral', '').strip()

        ADMIN_CODE = 'RAHASIAADMIN2025' # Referral untuk admin

        if form.is_valid():
            user = form.save(commit=False)

            if referral_code == ADMIN_CODE:
                user.is_staff = True
                user.is_superuser = True
                role_msg = 'Admin'
            else:
                user.is_staff = False
                user.is_superuser = False
                role_msg = 'Annotator'

            user.save()
            messages.success(request, f'Akun berhasil dibuat sebagai {role_msg}!')
            return redirect('main:login')
    else:
        form = CustomUserCreationForm()

    return render(request, 'register.html', {'form': form})


def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('main:landing_page')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('main:login')

"""PASSWORD MANAGEMENT"""

def forgot_password(request):
    if request.method == 'POST':
        form = CustomPasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            user = User.objects.get(email=email)
            
            # Generate token dan uid
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            # Create reset URL
            current_site = get_current_site(request)
            reset_url = f"http://{current_site.domain}/reset-password/{uid}/{token}/"
            
            # Prepare email content
            subject = 'Reset Password - Sign Language Video Annotator'
            message = f'''
Halo {user.username},

Anda meminta reset password untuk akun Sign Language Video Annotator.
Klik link berikut untuk mengatur password baru:

{reset_url}

Link ini berlaku selama 24 jam.
Jika Anda tidak meminta reset password, abaikan email ini.

Terima kasih,
Tim Sign Language Video Annotator
            '''
            
            # Send email
            try:
                from django.conf import settings
                
                # Untuk development, selalu tampilkan di console
                if settings.DEBUG:
                    print("=" * 60)
                    print("🔑 RESET PASSWORD EMAIL")
                    print("=" * 60)
                    print(f"To: {email}")
                    print(f"Subject: {subject}")
                    print("-" * 60)
                    print(message)
                    print("=" * 60)
                    print(f"Direct Link: {reset_url}")
                    print("=" * 60)
                
                send_mail(
                    subject, 
                    message, 
                    settings.DEFAULT_FROM_EMAIL, 
                    [email],
                    fail_silently=False
                )
                
                if settings.DEBUG:
                    messages.success(
                        request, 
                        f'✅ Email reset password telah dikirim ke {email}. '
                        f'Periksa console untuk melihat email (mode development).'
                    )
                else:
                    messages.success(
                        request, 
                        f'Link reset password telah dikirim ke {email}. '
                        f'Silakan periksa inbox dan spam folder Anda.'
                    )
                
            except Exception as e:
                logger.error(f"Error sending reset email: {str(e)}")
                print(f"❌ Email Error: {str(e)}")
                print(f"🔗 Fallback - Reset URL: {reset_url}")
                
                messages.warning(
                    request, 
                    f'Email tidak dapat dikirim, tetapi Anda dapat menggunakan link berikut: '
                    f'<a href="{reset_url}" target="_blank">Reset Password</a>'
                )
            
            return redirect('main:login')
    else:
        form = CustomPasswordResetForm()
    
    return render(request, 'forgot_password.html', {'form': form})

def reset_password(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            form = CustomSetPasswordForm(user, request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Password berhasil diubah! Silakan login dengan password baru.')
                return redirect('main:login')
        else:
            form = CustomSetPasswordForm(user)
        
        return render(request, 'reset_password.html', {'form': form})
    else:
        messages.error(request, 'Link reset password tidak valid atau sudah kedaluwarsa.')
        return redirect('main:forgot_password')

@login_required
def change_password(request):
    if request.method == 'POST':
        form = CustomPasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)  # Agar user tidak logout
            messages.success(request, 'Password berhasil diubah!')
            return redirect('main:landing_page')
    else:
        form = CustomPasswordChangeForm(user=request.user)
    
    return render(request, 'change_password.html', {'form': form})

"""KHUSUS UNTUK PAGE"""

@csrf_exempt
@login_required
def landing_page(request):
    """ Menampilkan halaman landing page dengan folder-folder """
    all_videos = Video.objects.all()
    folder_dict = defaultdict(list)

    # Kelompokkan video berdasarkan folder_name
    for video in all_videos:
        folder_dict[video.folder_name].append(video)

    # Buat list folder dengan status annotasi
    folder_info = []
    for folder_name, videos in folder_dict.items():
        total = len(videos)
        annotated = sum(1 for v in videos if v.is_annotated)
        # Untuk display nama folder
        parts = folder_name.split('_')
        lembaga = parts[0]
        jenis = 'SIBI' if parts[1] == 'SB' else parts[1]
        tanggal_str = parts[2]  # e.g: 290120

        # Formating tanggal
        day = tanggal_str[0:2]
        month = tanggal_str[2:4]
        year = '20' + tanggal_str[4:6]
        tanggal_format = f"{day}/{month}/{year}"
        display_name = f'{lembaga} - {jenis} - {tanggal_format}'

        # display_name = format_folder_display(folder_name)
        all_done = (annotated == total)
        folder_info.append({
            'name': folder_name,
            'annotated': annotated,
            'total': total,
            'all_done': all_done,
            'display_name': display_name
        })

    return render(request, 'landing_page.html', {'folders': folder_info})

@csrf_exempt
@login_required
def folder_page(request, folder_name):
    """ Menampilkan halaman folder dengan daftar video di dalamnya """
    videos = Video.objects.filter(folder_name=folder_name)
    # display_name = format_folder_display(folder_name)
    
    return render(request, 'folder_page.html', {'folder_name': folder_name, 'videos': videos})

@csrf_exempt
@login_required
def video_editor_page(request, video_title):
    video = get_object_or_404(Video, title=video_title)

    context = {
        'video': video,
        'automated_transcript': video.automated_transcript,
        'transcript_alignment': video.transcript_alignment,
        'sibi_sentence': video.sibi_sentence,
        'potential_problem': video.potential_problem,
        'comment': video.comment,
        'annotated_by': video.annotated_by.username if video.annotated_by else '-',
        'is_annotated': video.is_annotated,
        'created_at': video.created_at,
    }

    return render(request, 'video_editor.html', context)

@csrf_exempt
@login_required
def upload_video_page(request):
    return render(request, 'upload_video.html')

@csrf_exempt
@login_required
def upload_transcript_page(request):
    return render(request, 'upload_transcript.html')

@csrf_exempt
@login_required
def upload_file_page(request):
    if not request.user.is_superuser:
        return redirect('main:landing_page')
    return render(request, 'upload_file.html')

# ENHANCED USER INTERFACE

@csrf_exempt
@login_required
def format_folder_display(name):
    """ Format nama folder untuk tampilan yang lebih user-friendly """
    try:
        parts = name.split('_')
        lembaga = parts[0]
        jenis = 'SIBI' if parts[1] == 'SB' else parts[1]
        tanggal_str = parts[2]  # e.g: 290120

        # Formating tanggal
        day = tanggal_str[0:2]
        month = tanggal_str[2:4]
        year = '20' + tanggal_str[4:6]
        tanggal_format = f"{day}/{month}/{year}"

        return f'{lembaga} - {jenis} - {tanggal_format}'
    except Exception:
        return name # Kalo misalnya format namanya ga sesuai

@csrf_exempt
@login_required
def landing_page_data(request):
    """ Menampilkan halaman landing page dengan folder-folder """
    all_videos = Video.objects.all()
    folder_dict = defaultdict(list)

    # Kelompokkan video berdasarkan folder_name
    for video in all_videos:
        folder_dict[video.folder_name].append(video)

    # Buat list folder dengan status annotasi
    folder_info = []
    for folder_name, videos in folder_dict.items():
        total = len(videos)
        annotated = sum(1 for v in videos if v.is_annotated)
        # display_name = format_folder_display(str(folder_name))
        all_done = (annotated == total)
        folder_info.append({
            'name': folder_name,
            'annotated': annotated,
            'total': total,
            'all_done': all_done,
            # 'display_name': display_name,
        })

    return JsonResponse({
        'folders': [
            {
                'name': folder['name'],
                'annotated': folder['annotated'],
                'total': folder['total'],
                'all_done': folder['all_done'],
                # 'display_name': folder['display_name'],
            } for folder in folder_info
        ]
    })