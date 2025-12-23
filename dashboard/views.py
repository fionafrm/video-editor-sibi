from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
import io
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from main.models import Video
from .models import UserProfile
from .forms import UserEditForm, UserProfileForm

# Create your views here.
def format_folder_display(folder_name):
    """Format folder name untuk display (e.g., TVRI_SB_061119 -> TVRI - SIBI - 06/11/19)"""
    try:
        parts = folder_name.split('_')
        lembaga = parts[0]
        jenis = 'SIBI' if parts[1] == 'SB' else parts[1]
        tanggal_str = parts[2]  # e.g: 290120
        
        # Formating tanggal
        day = tanggal_str[0:2]
        month = tanggal_str[2:4]
        year = '20' + tanggal_str[4:6]
        tanggal_format = f"{day}/{month}/{year}"
        display_name = f'{lembaga} - {jenis} - {tanggal_format}'
        
        return display_name
    except:
        # Jika format tidak sesuai, kembalikan nama asli
        return folder_name

@login_required
def dashboard(request):
    user = request.user
    
    # Total video count
    total_videos = Video.objects.count()
    
    # Total annotations count (global)
    total_annotations = Video.objects.filter(is_annotated=True).count()
    
    # User's total annotations
    user_total_annotations = Video.objects.filter(annotated_by=user, is_annotated=True).count()
    
    # Get user's 3 most recent annotated videos (videos accessed by this user)
    user_recent_videos = Video.objects.filter(annotated_by=user).order_by('-created_at')[:3]
    
    # Get available folders for export options with display names
    folder_names = Video.objects.exclude(folder_name__isnull=True).exclude(folder_name='').values_list('folder_name', flat=True).distinct().order_by('folder_name')
    
    # Create folder list with both original name and display name
    folders = []
    for folder_name in folder_names:
        folders.append({
            'name': folder_name,
            'display_name': format_folder_display(folder_name)
        })
    
    # Get user statistics for admin
    user_stats = []
    remaining_videos = 0
    
    if user.is_superuser:
        from django.contrib.auth.models import User
        from django.db.models import Count, Q
        
        users = User.objects.all().annotate(
            annotation_count=Count('video', filter=Q(video__is_annotated=True))
        ).order_by('-annotation_count')
        
        for u in users:
            full_name = u.get_full_name() if u.get_full_name() else f"{u.first_name} {u.last_name}".strip()
            user_stats.append({
                'username': u.username,
                'full_name': full_name,
                'annotation_count': u.annotation_count,
                'is_superuser': u.is_superuser
            })
    else:
        # Untuk user biasa, hitung statistik personal mereka
        total_user_videos = Video.objects.count()  # Total video yang available untuk semua user
        remaining_videos = total_user_videos - user_total_annotations
    
    context = {
        'total_videos': total_videos,
        'total_annotations': total_annotations,
        'user_total_annotations': user_total_annotations,
        'remaining_videos': remaining_videos,
        'videos': user_recent_videos,
        'folders': folders,
        'user_stats': user_stats
    }
    
    return render(request, "dashboard.html", context)


@login_required
def profile(request):
    user = request.user
    
    # Ensure user has a profile
    profile, created = UserProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        # Handle form submission
        user_form = UserEditForm(request.POST, instance=user)
        profile_form = UserProfileForm(request.POST, instance=profile)
        
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'Profil berhasil diperbarui!')
            return redirect('dashboard:profile')
        else:
            messages.error(request, 'Terjadi kesalahan. Silakan periksa form kembali.')
    else:
        # Handle GET request - display forms
        user_form = UserEditForm(instance=user)
        profile_form = UserProfileForm(instance=profile)
    
    # Calculate user-specific statistics
    user_projects = Video.objects.filter(annotated_by=user).count() if user.is_authenticated else 0
    user_annotations = Video.objects.filter(annotated_by=user, is_annotated=True).count() if user.is_authenticated else 0
    
    # Determine role
    if user.is_superuser:
        role = "Admin"
    else:
        role = "Video Annotator"
    
    context = {
        'user': user,
        'role': role,
        'user_projects': user_projects,
        'user_annotations': user_annotations,
        'user_form': user_form,
        'profile_form': profile_form,
        'profile': profile,
    }
    
    return render(request, "profile.html", context)


def create_excel_workbook(videos, sheet_name="Video Data"):
    """
    Create Excel workbook from video queryset
    """
    # Create workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    
    # Define headers
    headers = [
        'Title', 
        'Folder Name',
        'File Path',
        'Automated Transcript',
        'Transcript Alignment', 
        'SIBI Sentence',
        'Potential Problem',
        'Comment',
        'Is Annotated',
        'Annotated By',
        'Merged Video Path',
        'Created At'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers to worksheet
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_alignment
    
    # Style for data cells
    data_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Add video data
    for row, video in enumerate(videos, 2):  # Start from row 2 (after headers)
        data = [
            video.title,
            video.folder_name or '',
            video.file.name if video.file else '',
            video.automated_transcript or '',
            video.transcript_alignment or '',
            video.sibi_sentence or '',
            video.potential_problem or '',
            video.comment or '',
            'Yes' if video.is_annotated else 'No',
            video.annotated_by.username if video.annotated_by else '',
            video.merged_video_path or '',
            video.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        for col, value in enumerate(data, 1):
            cell = worksheet.cell(row=row, column=col, value=value)
            cell.border = data_border
            cell.alignment = data_alignment
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = max(len(headers[col-1]), 15)  # Minimum width
        
        # Check data length (sample first 100 rows for performance)
        for row in range(2, min(102, worksheet.max_row + 1)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value:
                cell_length = len(str(cell_value)[:100])
                max_length = max(max_length, cell_length)
        
        # Set column width (max 50 characters for readability)
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return workbook


@login_required
def export_videos(request):
    """
    Export video data to Excel - supports filtering by folder
    """
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({
            'error': 'openpyxl library is not installed. Please install it to use export feature.'
        }, status=500)
    
    try:
        # Get folder filter from request
        folder_name = request.GET.get('folder', None)
        
        # Filter videos
        if folder_name and folder_name != 'all':
            videos = Video.objects.filter(folder_name=folder_name).order_by('-created_at')
            sheet_name = f"Videos - {folder_name}"
            filename_part = folder_name.replace('/', '_').replace('\\', '_')
        else:
            videos = Video.objects.all().order_by('-created_at')
            sheet_name = "All Videos"
            filename_part = "all"
        
        # Create workbook
        workbook = create_excel_workbook(videos, sheet_name)
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'video_data_{filename_part}_{timestamp}.xlsx'
        
        # Create in-memory file
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        # Return error response
        return JsonResponse({
            'error': f'Error exporting data: {str(e)}'
        }, status=500)